import pandas as pd
import googlemaps
import json
import os
from tkinter import Tk, Label, Entry, Button, filedialog, messagebox, Text, END, Checkbutton, BooleanVar
from openpyxl.utils import get_column_letter
from urllib.parse import quote

# Percorso del file con l'API
CONFIG_FILE = "apikey.json"

# Funzione per generare link Google Maps distanza in macchina, per qualche ragione i risultati diretti dell'API si discostano da quelli su sito Gmaps
def genera_link_maps(origine, destinazione):
    base_url = "https://www.google.com/maps/dir/?api=1"
    params = {
        'origin': quote(origine),
        'destination': quote(destinazione),
        'travelmode': 'driving'
    }
    return f"{base_url}&origin={params['origin']}&destination={params['destination']}&travelmode={params['travelmode']}"

# Funzione per caricare l'API Key salvata
def carica_api_key():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r") as file:
            config = json.load(file)
            return config.get("api_key", "")
    return ""

# Funzione per salvare l'API Key
def salva_api_key(api_key):
    with open(CONFIG_FILE, "w") as file:
        json.dump({"api_key": api_key}, file)

# Funzione per calcolare la distanza a piedi in km
def calcola_distanza_a_piedi(origine, destinazione):
    try:
        # Ottiene le indicazioni stradali a piedi
        direzioni = gmaps.directions(
            origine,
            destinazione,
            mode="walking",
            departure_time=None
        )

        # Estrai la distanza in metri e convertila in km
        distanza_metri = direzioni[0]['legs'][0]['distance']['value']
        distanza_km = distanza_metri / 1000
        return round(distanza_km, 2)  # Arrotonda a 2 decimali
    except Exception as e:
        print(f"Errore nel calcolo della distanza a piedi tra {origine} e {destinazione}: {e}")
        return None

# Funzione per calcolare il tempo in macchina
def calcola_tempo_in_macchina(origine, destinazione):
    try:
        # Ottieni le indicazioni stradali in macchina
        direzioni = gmaps.directions(
            origine,
            destinazione,
            mode="driving",
            departure_time=None
        )

        # Estrai il tempo di percorrenza
        tempo = direzioni[0]['legs'][0]['duration']['text']
        return tempo
    except Exception as e:
        print(f"Errore nel calcolo del tempo in macchina tra {origine} e {destinazione}: {e}")
        return "Errore"

# Funzione principale per avviare il calcolo
def avvia_calcolo():
    # Ottieni i valori inseriti dall'utente
    api_key = entry_api_key.get()
    file_excel = entry_file_excel.get()
    riferimento = entry_riferimento.get()
    cartella_output = entry_cartella_output.get()
    nome_file_output = entry_nome_file_output.get()

    # Salva l'API Key se la casella è spuntata
    if salva_api_key_var.get():
        salva_api_key(api_key)

    # Verifica che tutti i campi siano stati inseriti
    if not api_key or not file_excel or not riferimento or not cartella_output or not nome_file_output:
        messagebox.showerror("Errore", "Tutti i campi sono obbligatori!")
        return

    # Inizializza il client di Google Maps
    global gmaps
    gmaps = googlemaps.Client(key=api_key)

    # Carica il file Excel
    try:
        df = pd.read_excel(file_excel)
    except Exception as e:
        messagebox.showerror("Errore", f"Impossibile leggere il file Excel: {e}")
        return

    # Trova la colonna "indirizzo" (case not sensitive)
    colonna_indirizzo = None
    for col in df.columns:
        if col.lower() == "indirizzo":
            colonna_indirizzo = col
            break

    if not colonna_indirizzo:
        messagebox.showerror("Errore", "La colonna 'indirizzo' non esiste nel file Excel.")
        return

    # Crea nuove colonne per i risultati
    colonna_distanza = f"Distanza A PIEDI da {riferimento} (km)"
    colonna_tempo = "Stima tempo IN MACCHINA "
    colonna_link = "Link Google Maps per verifica manuale distanza IN MACCHINA"

    df[colonna_distanza] = None
    df[colonna_tempo] = None
    df[colonna_link] = None

    # Itera sulle righe e calcola le distanze e i tempi
    for index, row in df.iterrows():
        indirizzo = row[colonna_indirizzo]

        # Verifica che l'indirizzo sia presente
        if pd.notna(indirizzo):
            # Calcola la distanza a piedi
            distanza = calcola_distanza_a_piedi(indirizzo, riferimento)
            if distanza is not None:
                df.loc[index, colonna_distanza] = distanza
                log_text.insert(END, f"Distanza a piedi tra {indirizzo} e {riferimento}: {distanza} km\n")
            else:
                log_text.insert(END, f"Errore nel calcolo della distanza a piedi per la riga {index + 1}\n")

            # Calcola il tempo in macchina
            tempo = calcola_tempo_in_macchina(indirizzo, riferimento)
            if tempo != "Errore":
                df.loc[index, colonna_tempo] = tempo
                log_text.insert(END, f"Tempo in macchina tra {indirizzo} e {riferimento}: {tempo}\n")
            else:
                log_text.insert(END, f"Errore nel calcolo del tempo in macchina per la riga {index + 1}\n")

            # Genera link Google Maps
            link = genera_link_maps(indirizzo, riferimento)
            df.loc[index, colonna_link] = link
            log_text.insert(END, f"Link generato per {indirizzo} -> {riferimento}\n")
        else:
            log_text.insert(END, f"Indirizzo mancante alla riga {index + 1}\n")

    # Ordina le righe in base alla distanza crescente
    df.sort_values(by=colonna_distanza, inplace=True)

    # Salva i risultati in un nuovo file Excel
    output_file = f"{cartella_output}/{nome_file_output}.xlsx"
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

        # Imposta la larghezza delle colonne
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        worksheet.column_dimensions[get_column_letter(df.columns.get_loc(colonna_distanza) + 1)].width = 80
        worksheet.column_dimensions[get_column_letter(df.columns.get_loc(colonna_tempo) + 1)].width = 30
        worksheet.column_dimensions[get_column_letter(df.columns.get_loc(colonna_link) + 1)].width = 60

    messagebox.showinfo("Calcolo completato",
                        f"Calcolo delle distanze e dei tempi completato! I risultati sono stati salvati in {output_file}")

# Funzioni per la selezione file/directory
def seleziona_file(entry_widget):
    percorso = filedialog.askopenfilename()
    if percorso:
        entry_widget.delete(0, END)
        entry_widget.insert(0, percorso)

def seleziona_cartella(entry_widget):
    percorso = filedialog.askdirectory()
    if percorso:
        entry_widget.delete(0, END)
        entry_widget.insert(0, percorso)

# Creazione della finestra principale
root = Tk()
root.title("Calcolo Distanze e Tempi con Google Maps")

# Breve descrizione del programma
descrizione = (
    "Questo programma calcola la distanza a piedi e dà una stima del tempo in macchina "
    "tra un indirizzo di riferimento e una lista di indirizzi presenti in un file Excel."
    "\nInfine fornisce un'ultima colonna di link diretti a Google Maps per la verifica manuale\n\n"
    "Istruzioni:\n"
    "1. Inserisci la tua API Key di Google Maps.\n"
    "2. Seleziona il file Excel contenente gli indirizzi in una colonna chiamata 'indirizzo'.\n"
    "3. Inserisci l'indirizzo di riferimento (formato ideale: via, numero civico, cap, città).\n"
    "4. Specifica la cartella di output e il nome del file Excel dei risultati.\n"
    "5. Clicca su 'Avvia Calcolo'.\n\n"

    "Note:\nIl programma ha bisogno di almeno via, numero civico e cap per funzionare adeguatamente"
    "\nSu excel si possono fondere le colonne di indirizzo e cap con la funzione =A1 & \", \" & B1 "
    "\nSarà poi importante copiare i risultati come valori su una nuova colonna 'indirizzo'"
    "\nVerrà creato un file apikey.json per salvare la key google"
)

Label(root, text=descrizione, justify="left", wraplength=500).grid(row=0, column=0, columnspan=3, padx=10, pady=10)

# Etichette e campi di input
Label(root, text="API Key di Google Maps:").grid(row=1, column=0, padx=10, pady=10)
entry_api_key = Entry(root, width=50)
entry_api_key.grid(row=1, column=1, padx=10, pady=10)

# Carica l'API Key salvata, se presente
entry_api_key.insert(0, carica_api_key())

# Spunta per salvare l'API Key
salva_api_key_var = BooleanVar()
Checkbutton(root, text="Salva API Key", variable=salva_api_key_var).grid(row=1, column=2, padx=10, pady=10)

Label(root, text="File Excel:").grid(row=2, column=0, padx=10, pady=10)
entry_file_excel = Entry(root, width=50)
entry_file_excel.grid(row=2, column=1, padx=10, pady=10)
Button(root, text="Sfoglia", command=lambda: seleziona_file(entry_file_excel)).grid(row=2, column=2, padx=10, pady=10)

Label(root, text="Indirizzo di Riferimento:").grid(row=3, column=0, padx=10, pady=10)
entry_riferimento = Entry(root, width=50)
entry_riferimento.grid(row=3, column=1, padx=10, pady=10)

Label(root, text="Cartella di Output:").grid(row=4, column=0, padx=10, pady=10)
entry_cartella_output = Entry(root, width=50)
entry_cartella_output.grid(row=4, column=1, padx=10, pady=10)
Button(root, text="Sfoglia", command=lambda: seleziona_cartella(entry_cartella_output)).grid(row=4, column=2, padx=10, pady=10)

Label(root, text="Nome File Output:").grid(row=5, column=0, padx=10, pady=10)
entry_nome_file_output = Entry(root, width=50)
entry_nome_file_output.grid(row=5, column=1, padx=10, pady=10)

# Area di log(superflua?)
Label(root, text="Log:").grid(row=6, column=0, padx=10, pady=10)
log_text = Text(root, height=10, width=70)
log_text.grid(row=6, column=1, columnspan=2, padx=10, pady=10)

# Pulsante per avviare il calcolo
Button(root, text="Avvia Calcolo", command=avvia_calcolo).grid(row=7, column=1, padx=10, pady=20)

# Avvio dell'interfaccia grafica
root.mainloop()
